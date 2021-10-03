from bs4 import BeautifulSoup
import urllib.request as req
import openpyxl
import requests
import os
from openpyxl.drawing.image import Image


# 엑셀 파일 존재하는지 확인
if not os.path.exists("./진아멜론차트.xlsx"):
    book = openpyxl.Workbook()
    book.save("./진아멜론차트.xlsx")

# 서버에게 허락맡기
header = req.Request("https://www.melon.com/chart/index.htm", headers={"User-Agent":"Mozilla/5.0"})

# url = 'https://www.melon.com/chart/index.htm'
# headers2 = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36'}
# r = requests.get(url, headers=headers2)
# soup = BeautifulSoup(r.text, 'html.parser')

code = req.urlopen(header)
soup = BeautifulSoup(code, "html.parser")
title = soup.select("div.ellipsis.rank01 > span > a") # 속성값에 띄어쓰기가 있으면 해당 공백은 . 으로 표시한다
name = soup.select("div.ellipsis.rank02 > span.checkEllipsis")
album = soup.select("div.ellipsis.rank03 > a")
image = soup.select("a.image_typeAll > img")

'''
    openpyxl
    - 내가 기록한 excel 파일을 가져와야 한다 
    - 엑셀에서 할 수 있는 모든 기능을 파이썬에서 거의 할 수 있다
    - 모든 것은 디테일하게 말해줘야 한다
    - 어떤 sheet를 사용할 것인지도 말해줘야 한다
'''
book = openpyxl.load_workbook("./진아멜론차트.xlsx")
sheet = book.active # 자동으로 열리는 그 시트를 사용하겠다는 의미
sheet.column_dimensions["A"].width = "15"
sheet.column_dimensions["B"].width = "50"
sheet.column_dimensions["C"].width = "30"
sheet.column_dimensions["D"].width = "35"

row_num = 1
column_num = 2

# 이미지 저장 폴더 만들기
if not os.path.exists("./진아멜론차트"):
    os.mkdir("./진아멜론차트")

for i in range(len(title)):

    req.urlretrieve(image[i].attrs["src"], "./진아멜론차트/{}.png".format(row_num))

    print(title[i].string, name[i].text, album[i].string, image[i].attrs["src"])

    img_for_excel = Image("./진아멜론차트/{}.png".format(row_num)) # excel 에 넣을 수 있는 형태로 이미지 변환

    sheet.row_dimensions[row_num].height = 95

    sheet.add_image(img_for_excel, "A{}".format(row_num))

    sheet.cell(row=row_num, column=2).value = title[i].string # 엑셀 파일의 cell 클릭
    sheet.cell(row=row_num, column=3).value = name[i].text
    sheet.cell(row=row_num, column=4).value = album[i].string
    book.save("./진아멜론차트.xlsx")
    row_num += 1

'''
    - Error 안 나는 방법으로 출력하면 됨
    
    .string
    .text
'''

'''
    HTTP Error 406 : 서버가 우리의 접속을 차단
    
    - 서버는 요청 메시지를 확인할 때, 악의적인 의도를 가지고 접속하는지 확인한다. (hacker or crolling bot) 
'''
